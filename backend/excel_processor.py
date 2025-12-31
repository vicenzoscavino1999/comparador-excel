import pandas as pd
import io
import re
from typing import Tuple, Optional, List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# Column detection patterns - expanded for better matching
CODIGO_PATTERNS = [
    r'^c[oó]digo.*$', r'^cod\.?.*$', r'^sku.*$', r'^id$', r'^codigo$',
    r'^c[oó]d\.?\s*producto.*$', r'^item$', r'^referencia.*$', r'^ref\.?$',
    r'^art[ií]culo.*$', r'^cod.*art.*$', r'^clave.*$', r'^num.*$',
    r'^n[uú]mero.*$', r'^parte.*$', r'^c[oó]d.*$', r'^barcode.*$',
    r'^upc$', r'^ean$', r'^plu$'  # Removed 'material' - too ambiguous
]

PRODUCTO_PATTERNS = [
    r'^producto.*$', r'^descripci[oó]n.*$', r'^nombre.*$', r'^art[ií]culo.*$',
    r'^item$', r'^detalle.*$', r'^material.*$', r'^desc\.?.*$',
    r'^denominaci[oó]n.*$', r'^especificaci[oó]n.*$', r'^concepto.*$'
]

CANTIDAD_PATTERNS = [
    r'^cant\.?\s*final.*$', r'^cantidad.*$', r'^cant\.?.*$', r'^qty.*$', r'^unidades.*$', r'^stock.*$',
    r'^existencia.*$', r'^saldo.*$', r'^und\.?.*$', r'^pzs\.?.*$',
    r'^total.*$', r'^unid.*$', r'^piezas.*$', r'^disponible.*$',
    r'^inventario.*$', r'^disp.*$', r'^almac[eé]n.*$', r'^bodega.*$',
    r'^f[ií]sico.*$', r'^conteo.*$'
]


def detect_column(df: pd.DataFrame, patterns: List[str]) -> Optional[str]:
    """Detect a column matching any of the given patterns"""
    for col in df.columns:
        col_clean = str(col).lower().strip()
        for pattern in patterns:
            if re.match(pattern, col_clean, re.IGNORECASE):
                return col
    return None


def read_excel_file(file_content: bytes, filename: str) -> pd.DataFrame:
    """Read Excel file (supports both .xls and .xlsx) with all columns as text to preserve leading zeros.
    
    Optimized: Only reads first 30 rows initially for header detection to reduce memory usage.
    """
    file_buffer = io.BytesIO(file_content)
    is_xls = filename.lower().endswith('.xls')
    engine = 'xlrd' if is_xls else 'openpyxl'
    
    # OPTIMIZATION: Read only first 30 rows for header detection (saves RAM for large files)
    header_sample = pd.read_excel(file_buffer, engine=engine, header=None, dtype=str, nrows=30)
    
    # Try to find the header row in the sample
    header_row = None
    header_keywords = ['codigo', 'código', 'cod', 'producto', 'descripcion', 'descripción', 
                      'cantidad', 'cant', 'stock', 'unidades', 'total', 'item', 'articulo',
                      'artículo', 'material', 'referencia', 'nombre', 'detalle']
    
    for i in range(len(header_sample)):
        row_values = [str(v).lower().strip() for v in header_sample.iloc[i].tolist()]
        matches = sum(1 for v in row_values for kw in header_keywords if kw in v.lower())
        if matches >= 2:  # At least 2 header keywords found
            header_row = i
            break
    
    # Now read the full file with the detected header
    file_buffer.seek(0)
    if header_row is not None:
        df = pd.read_excel(file_buffer, engine=engine, header=header_row, dtype=str)
    else:
        # No headers found, use positional columns
        df = pd.read_excel(file_buffer, engine=engine, header=None, dtype=str)
        df = assign_positional_columns(df)
    
    return df


def assign_positional_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Assign column names based on data patterns when no headers exist"""
    # Analyze each column to guess its purpose
    new_columns = []
    codigo_assigned = False
    cantidad_assigned = False
    
    for i, col in enumerate(df.columns):
        col_data = df[col].dropna()
        if len(col_data) == 0:
            new_columns.append(f'Columna_{i}')
            continue
            
        # Check if mostly numeric (could be código or cantidad)
        numeric_count = pd.to_numeric(col_data, errors='coerce').notna().sum()
        is_mostly_numeric = numeric_count / len(col_data) > 0.7
        
        # Check if looks like text descriptions (long strings)
        avg_length = col_data.astype(str).str.len().mean()
        
        if not codigo_assigned and is_mostly_numeric and avg_length < 15:
            new_columns.append('Código')
            codigo_assigned = True
        elif not cantidad_assigned and is_mostly_numeric and codigo_assigned:
            new_columns.append('Cantidad')
            cantidad_assigned = True
        elif avg_length > 10:  # Likely product description
            new_columns.append('Producto')
        else:
            new_columns.append(f'Columna_{i}')
    
    df.columns = new_columns
    return df


def clean_code_format(code: str) -> str:
    """Clean code format: remove .0 suffix from codes that Excel converted to float.
    
    Uses Decimal for precision with long codes (>15 digits) instead of float.
    """
    from decimal import Decimal, InvalidOperation
    
    code = str(code).strip().upper()
    # Remove .0 suffix (e.g., '806.0' -> '806')
    if code.endswith('.0'):
        code = code[:-2]
    # Handle scientific notation (e.g., '1.23E+10' -> full number) using Decimal for precision
    try:
        if 'E' in code.upper():
            # Use Decimal to preserve precision for long codes
            dec_val = Decimal(code)
            if dec_val == dec_val.to_integral_value():
                code = str(int(dec_val))
    except (InvalidOperation, ValueError, OverflowError):
        pass
    return code


def clean_quantity(value) -> float:
    """Clean quantity value: handle thousands separators, decimal variations, spaces"""
    if pd.isna(value) or value is None:
        return 0.0
    
    val_str = str(value).strip()
    if not val_str or val_str.upper() == 'NAN':
        return 0.0
    
    # Remove spaces
    val_str = val_str.replace(' ', '')
    
    # Handle different number formats:
    # "1,234.56" (US/UK) -> 1234.56
    # "1.234,56" (EU) -> 1234.56
    # "1,234" could be 1234 (thousands) or 1.234 (decimal)
    
    # Count occurrences of . and ,
    dots = val_str.count('.')
    commas = val_str.count(',')
    
    if dots == 1 and commas == 0:
        # Standard decimal: "1234.56"
        pass
    elif dots == 0 and commas == 1:
        # Could be "1,234" (thousands) or "1,5" (EU decimal)
        # If comma is in last 3 positions and only 1-2 digits after, treat as decimal
        comma_pos = val_str.rfind(',')
        after_comma = len(val_str) - comma_pos - 1
        if after_comma <= 2:
            # EU decimal format: "1,5" or "1,50"
            val_str = val_str.replace(',', '.')
        else:
            # Thousands separator: "1,234"
            val_str = val_str.replace(',', '')
    elif dots >= 1 and commas >= 1:
        # Mixed format
        # If last separator is comma, it's EU: "1.234,56"
        # If last separator is dot, it's US: "1,234.56"
        last_dot = val_str.rfind('.')
        last_comma = val_str.rfind(',')
        if last_comma > last_dot:
            # EU format: "1.234,56"
            val_str = val_str.replace('.', '').replace(',', '.')
        else:
            # US format: "1,234.56"
            val_str = val_str.replace(',', '')
    elif commas > 1:
        # Multiple commas = thousands separators: "1,234,567"
        val_str = val_str.replace(',', '')
    elif dots > 1:
        # Multiple dots = thousands separators (EU): "1.234.567"
        val_str = val_str.replace('.', '')
    
    try:
        return float(val_str)
    except ValueError:
        return 0.0


def clean_dataframe(df: pd.DataFrame, codigo_col: str, producto_col: Optional[str], cantidad_col: str) -> pd.DataFrame:
    """Clean the dataframe: remove duplicate headers, normalize data"""
    # Make a copy
    df = df.copy()
    
    # Remove rows that look like headers (where codigo column equals its column name)
    original_codigo_name = codigo_col
    mask = df[codigo_col].astype(str).str.lower().str.strip() != original_codigo_name.lower().strip()
    df = df[mask]
    
    # Normalize codigo: convert to string, clean format to preserve leading zeros
    df['Código'] = df[codigo_col].apply(clean_code_format)
    
    # Normalize cantidad: clean and convert to numeric (handles thousands separators, etc.)
    df['Cantidad'] = df[cantidad_col].apply(clean_quantity)
    
    # Handle producto
    if producto_col:
        df['Producto'] = df[producto_col].astype(str).str.strip()
        # Clean 'nan' strings
        df['Producto'] = df['Producto'].replace({'nan': '', 'NaN': '', 'NAN': ''})
    else:
        df['Producto'] = ''
    
    # Keep only relevant columns
    result = df[['Código', 'Producto', 'Cantidad']].copy()
    
    # Remove empty codes
    result = result[result['Código'] != '']
    result = result[result['Código'].str.upper() != 'NAN']
    
    return result


def aggregate_by_code(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate rows by code: sum quantities and take first non-empty product"""
    # Group by Código and aggregate
    def first_non_empty(series):
        for val in series:
            if val and str(val).strip() and str(val).strip().upper() != 'NAN':
                return val
        return ''
    
    aggregated = df.groupby('Código', as_index=False).agg({
        'Producto': first_non_empty,
        'Cantidad': 'sum'
    })
    return aggregated


def process_excel_file(file_content: bytes, filename: str) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Process an Excel file: read, detect columns, clean"""
    # Read file
    df = read_excel_file(file_content, filename)
    
    # Get list of actual columns for error messages
    available_cols = [str(col).strip() for col in df.columns.tolist()]
    available_cols_str = ", ".join(available_cols[:10])  # Show first 10
    if len(available_cols) > 10:
        available_cols_str += f"... (+{len(available_cols)-10} más)"
    
    # Detect columns
    codigo_col = detect_column(df, CODIGO_PATTERNS)
    producto_col = detect_column(df, PRODUCTO_PATTERNS)
    cantidad_col = detect_column(df, CANTIDAD_PATTERNS)
    
    if not codigo_col:
        raise ValueError(f"No se encontró la columna de Código. Columnas disponibles: {available_cols_str}")
    if not cantidad_col:
        raise ValueError(f"No se encontró la columna de Cantidad. Columnas disponibles: {available_cols_str}")
    
    detected = {
        "codigo": codigo_col,
        "producto": producto_col or "(no detectado)",
        "cantidad": cantidad_col
    }
    
    # Clean dataframe
    cleaned_df = clean_dataframe(df, codigo_col, producto_col, cantidad_col)
    
    # Aggregate duplicates by code (sum quantities, take first non-empty product)
    aggregated_df = aggregate_by_code(cleaned_df)
    
    return aggregated_df, detected


def compare_dataframes(df1: pd.DataFrame, df2: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """Compare two dataframes and generate comparison results"""
    # Rename columns for clarity
    df1 = df1.rename(columns={
        'Cantidad': 'Cantidad_Archivo1',
        'Producto': 'Producto_Archivo1'
    })
    df2 = df2.rename(columns={
        'Cantidad': 'Cantidad_Archivo2',
        'Producto': 'Producto_Archivo2'
    })
    
    # Full outer merge
    merged = pd.merge(
        df1, df2,
        on='Código',
        how='outer',
        indicator=True
    )
    
    # Fill NaN with 0 for quantities
    merged['Cantidad_Archivo1'] = merged['Cantidad_Archivo1'].fillna(0)
    merged['Cantidad_Archivo2'] = merged['Cantidad_Archivo2'].fillna(0)
    
    # Calculate difference
    merged['Diferencia'] = merged['Cantidad_Archivo1'] - merged['Cantidad_Archivo2']
    
    # Combine product names using coalesce (prefer file1, fallback to file2)
    merged['Producto'] = merged['Producto_Archivo1'].fillna('').replace('', pd.NA)
    merged['Producto'] = merged['Producto'].fillna(merged['Producto_Archivo2'].fillna(''))
    merged['Producto'] = merged['Producto'].fillna('').astype(str).str.strip()
    
    # Create different views
    comparison_complete = merged[['Código', 'Producto', 'Cantidad_Archivo1', 'Cantidad_Archivo2', 'Diferencia']].copy()
    
    only_differences = comparison_complete[comparison_complete['Diferencia'] != 0].copy()
    
    not_in_file2 = merged[merged['_merge'] == 'left_only'][['Código', 'Producto', 'Cantidad_Archivo1']].copy()
    not_in_file2 = not_in_file2.rename(columns={'Cantidad_Archivo1': 'Cantidad'})
    
    not_in_file1 = merged[merged['_merge'] == 'right_only'][['Código', 'Producto', 'Cantidad_Archivo2']].copy()
    not_in_file1 = not_in_file1.rename(columns={'Cantidad_Archivo2': 'Cantidad'})
    
    return {
        'comparison_complete': comparison_complete,
        'only_differences': only_differences,
        'not_in_file2': not_in_file2,
        'not_in_file1': not_in_file1
    }


def create_summary_data(df1: pd.DataFrame, df2: pd.DataFrame, results: Dict[str, pd.DataFrame]) -> List[List[Any]]:
    """Create summary statistics"""
    return [
        ['RESUMEN DE COMPARACIÓN', ''],
        ['', ''],
        ['Estadísticas Archivo 1', ''],
        ['Total de registros', len(df1)],
        ['Suma de cantidades', df1['Cantidad'].sum()],
        ['', ''],
        ['Estadísticas Archivo 2', ''],
        ['Total de registros', len(df2)],
        ['Suma de cantidades', df2['Cantidad'].sum()],
        ['', ''],
        ['Resultados de Comparación', ''],
        ['Total registros comparados', len(results['comparison_complete'])],
        ['Registros con diferencias', len(results['only_differences'])],
        ['Solo en Archivo 1', len(results['not_in_file2'])],
        ['Solo en Archivo 2', len(results['not_in_file1'])],
        ['', ''],
        ['Diferencia total de cantidades', results['comparison_complete']['Diferencia'].sum()],
    ]


def apply_styles(ws, df: pd.DataFrame, is_summary: bool = False):
    """Apply styles to a worksheet"""
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    
    positive_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    negative_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if is_summary:
        # Style summary sheet
        ws['A1'].font = Font(bold=True, size=14)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
                if cell.row in [3, 7, 11]:  # Section headers
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
    else:
        # Style data sheets
        # Header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows
        diff_col = None
        for idx, col in enumerate(ws[1], 1):
            if col.value == 'Diferencia':
                diff_col = idx
                break
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
                
                # Highlight differences
                if diff_col and cell.column == diff_col:
                    try:
                        val = float(cell.value) if cell.value else 0
                        if val > 0:
                            cell.fill = positive_fill
                        elif val < 0:
                            cell.fill = negative_fill
                    except (ValueError, TypeError):
                        pass
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def create_output_excel(df1: pd.DataFrame, df2: pd.DataFrame, results: Dict[str, pd.DataFrame]) -> bytes:
    """Create the output Excel file with all sheets"""
    wb = Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    summary_data = create_summary_data(df1, df2, results)
    for row in summary_data:
        ws_summary.append(row)
    apply_styles(ws_summary, None, is_summary=True)
    
    # Sheet 2: Complete Comparison
    ws_complete = wb.create_sheet("Comparación Completa")
    for r_idx, row in enumerate(dataframe_to_rows(results['comparison_complete'], index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_complete.cell(row=r_idx, column=c_idx, value=value)
    apply_styles(ws_complete, results['comparison_complete'])
    
    # Sheet 3: Only Differences
    ws_diff = wb.create_sheet("Solo Diferencias")
    if len(results['only_differences']) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(results['only_differences'], index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_diff.cell(row=r_idx, column=c_idx, value=value)
        apply_styles(ws_diff, results['only_differences'])
    else:
        ws_diff.cell(row=1, column=1, value="No hay diferencias entre los archivos")
    
    # Sheet 4: Not in File 2
    ws_not2 = wb.create_sheet("No Coinciden Archivo1")
    if len(results['not_in_file2']) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(results['not_in_file2'], index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_not2.cell(row=r_idx, column=c_idx, value=value)
        apply_styles(ws_not2, results['not_in_file2'])
    else:
        ws_not2.cell(row=1, column=1, value="Todos los códigos del Archivo 1 están en el Archivo 2")
    
    # Sheet 5: Not in File 1
    ws_not1 = wb.create_sheet("No Coinciden Archivo2")
    if len(results['not_in_file1']) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(results['not_in_file1'], index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_not1.cell(row=r_idx, column=c_idx, value=value)
        apply_styles(ws_not1, results['not_in_file1'])
    else:
        ws_not1.cell(row=1, column=1, value="Todos los códigos del Archivo 2 están en el Archivo 1")
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def process_comparison(file1_content: bytes, file1_name: str, 
                       file2_content: bytes, file2_name: str) -> Tuple[bytes, Dict[str, Any]]:
    """Main function to process the comparison"""
    # Process file 1
    df1, detected1 = process_excel_file(file1_content, file1_name)
    
    # Process file 2
    df2, detected2 = process_excel_file(file2_content, file2_name)
    
    # Compare
    results = compare_dataframes(df1, df2)
    
    # Create output
    output_bytes = create_output_excel(df1, df2, results)
    
    # Summary info
    info = {
        "file1": {
            "name": file1_name,
            "records": len(df1),
            "detected_columns": detected1
        },
        "file2": {
            "name": file2_name,
            "records": len(df2),
            "detected_columns": detected2
        },
        "results": {
            "total_compared": len(results['comparison_complete']),
            "with_differences": len(results['only_differences']),
            "only_in_file1": len(results['not_in_file2']),
            "only_in_file2": len(results['not_in_file1'])
        }
    }
    
    return output_bytes, info
